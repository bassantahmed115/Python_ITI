import openpyxl
import pandas as pd

def read_file(path, option):
    with open(path, option) as file:
        content = file.read()
    return content

def write_file(path, content):
    with open(path, 'w') as file:
        file.write(content)

def read_excel(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data

def write_excel(path, data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in data:
        sheet.append(row)
    workbook.save(path)
